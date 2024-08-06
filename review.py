import openpyxl
import datetime
import pickle


def review(user_name, learn_words='5'):
    learn_words = int(learn_words)
    res = []
    wave = '~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~'  #32
    file_name = 'record-'+user_name+'.xlsx'
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active

    today = datetime.date.today()
    # TODO time zone
    next_date_one = today + datetime.timedelta(days=1)
    today_str = today.strftime('%Y-%m-%d')
    next_date_one = next_date_one.strftime('%Y-%m-%d')

    diff = [1,1,2,3,3,5,5,10,10,20]
    row = 2
    new_word = 1
    new_word_list = []
    for entry in sheet.iter_rows(min_row=2,values_only=True):
        if entry[3] == None:
            if new_word > learn_words:
                for word in new_word_list:
                    res.append(word[0]+'\n')
                    res.append(word[1]+'\n')
                break
            res.append(wave+'\n'+entry[0]+'\n\n'+entry[1]+'\n'+wave+'\n')
            sheet.cell(row=row, column=3).value = today_str
            sheet.cell(row=row, column=4).value = next_date_one
            sheet.cell(row=row, column=5).value = '1'
            new_word += 1
            new_word_list.append([entry[0],entry[1]])
        elif entry[3] <= today_str:
            res.append(entry[0]+'\n')
            res.append(entry[1]+'\n')
            times = int(entry[4])
            next_date = today + datetime.timedelta(days=diff[times])
            next_date = next_date.strftime('%Y-%m-%d')
            sheet.cell(row=row, column=4).value = next_date
            sheet.cell(row=row, column=5).value = str(times+1)
        row += 1

    days = sheet.cell(row=2, column=8).value
    if sheet.cell(row=2, column=9).value != today_str:
        new_days = int(days)+1
        sheet.cell(row=2, column=8).value = str(new_days)
        sheet.cell(row=2, column=9).value = today_str
    else:
        new_days = days

    workbook.save(file_name)
    workbook.close()
    txt_name = 'today_words-'+user_name+'.txt'
    
    with open(txt_name, 'wb') as file:
        pickle.dump(res, file)
        
    return res, new_days

def get_review(user_name):
    file_name = 'record-'+user_name+'.xlsx'
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    today = datetime.date.today()
    today_str = today.strftime('%Y-%m-%d')
    if sheet.cell(row=2, column=9).value == today_str:
        res = []
        txt_name = 'today_words-'+user_name+'.txt'
        with open(txt_name, 'rb') as file:
            res = pickle.load(file)
        days = sheet.cell(row=2, column=8).value
    else:
        res = 'Learn it once!'
        days = '0'
    workbook.close()
    return res, days


def report_mistake(mistake_list, user_name):
    tomorrow = (datetime.date.today() + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
    file_name = 'record-'+user_name+'.xlsx'
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    search_string = mistake_list[0][:-1]
    count_list = 1
    count_row = 2
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == search_string:
            sheet.cell(row=count_row, column=4).value = tomorrow
            search_string = mistake_list[count_list][:-1]
            count_list += 1
            if count_list == len(mistake_list):
                break
        count_row += 1
        
    workbook.save(file_name)
    workbook.close()
    